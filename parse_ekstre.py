"""
parse_ekstre.py
───────────────
Ziraat Katılım Bankası hesap hareketi Excel dosyasını okur,
açıklama sütununu regex ile ayrıştırır ve yapılandırılmış
sütunlara sahip yeni bir Excel dosyası üretir.

KULLANIM:
    python parse_ekstre.py

ÇIKTI:
    Hesap_Hareketleri_PARSED.xlsx  (aynı klasörde)
"""

import re
import pandas as pd
from pathlib import Path

# ─────────────────────────────────────────
# 0. AYARLAR
# ─────────────────────────────────────────
GIRIS_DOSYA  = r"C:\Users\HUAWEİ\Downloads\Hesap_Hareketleri_22062026.xlsx"
CIKIS_DOSYA  = r"C:\Users\HUAWEİ\Downloads\Hesap_Hareketleri_PARSED.xlsx"


# ─────────────────────────────────────────
# 1. VERİYİ OKU
# ─────────────────────────────────────────
def veri_oku(dosya: str) -> pd.DataFrame:
    """Başlıksız Excel dosyasını okur, standart kolon adları atar."""
    df = pd.read_excel(dosya, header=None, dtype=str, keep_default_na=False)
    df.columns = ["islem_tarihi", "fis_no", "aciklama_ham", "yon_kod"]
    # Yön: 100 → GİRİŞ, -100 → ÇIKIŞ
    df["islem_yonu"] = df["yon_kod"].str.strip().apply(
        lambda x: "GİRİŞ" if x == "100" else ("ÇIKIŞ" if x == "-100" else x)
    )
    return df


# ─────────────────────────────────────────
# 2. IBAN FORMATLAYICI
# ─────────────────────────────────────────
def iban_formatla(iban_ham: str) -> str:
    """TR... formatındaki IBAN'ı TR00 0000 0000 ... şeklinde düzenler."""
    if not iban_ham:
        return ""
    iban = re.sub(r"\s+", "", iban_ham.upper())          # boşlukları kaldır
    if not re.match(r"^TR\d{24}$", iban):
        return iban_ham.strip()                           # Tanınmıyorsa ham hali
    parcalar = [iban[i:i+4] for i in range(0, len(iban), 4)]
    return " ".join(parcalar)


# ─────────────────────────────────────────
# 3. FORMAT PARSERları
# ─────────────────────────────────────────

def parse_GKDA(aciklama: str) -> dict | None:
    """
    FORMAT 1 — Döviz Alım
    Örnek: 0145GKDA25002033 Ref, Borçlu Müşteri: USD 140.000,00 Kur: 39,64 Döviz Alış İşlemi
    """
    p = re.compile(
        r"(0145GKDA\S+)\s+Ref,\s*Borçlu\s+Müşteri:\s*"
        r"([A-Z]{3})\s+([\d.,]+)\s+Kur:\s*([\d.,]+)",
        re.IGNORECASE
    )
    m = p.search(aciklama)
    if not m:
        return None
    return {
        "islem_tipi"      : "DOVIZ_ALIM",
        "banka_ref_no"    : m.group(1),
        "doviz_cinsi"     : m.group(2).upper(),
        "doviz_tutari"    : m.group(3).replace(".", "").replace(",", "."),
        "kur"             : m.group(4).replace(",", "."),
        "karsitaraf_ad"   : "",
        "karsitaraf_iban" : "",
        "karsitaraf_banka": "Ziraat Katılım Bankası",
        "aciklama_temiz"  : "Döviz Alış İşlemi",
        "cari_durumu"     : "BANKA_GİDERİ",
        "muhasebe_hesabi" : "102",
    }


def parse_HOHH(aciklama: str) -> dict | None:
    """
    FORMAT 2 — Havale Komisyonu / BSMV
    Örnek: 0145HOHH25238389 Referanslı Havale - BSMV Tahsilatı
           0145HOHH25238389 Referanslı Havale Komisyonu Tahsilatı
    """
    p = re.compile(r"(0145HOHH\S+)\s+(.+)", re.IGNORECASE)
    m = p.search(aciklama)
    if not m:
        return None
    acik = m.group(2).strip()
    tip  = "HAVALE_BSMV" if "BSMV" in acik.upper() else "HAVALE_KOM"
    return {
        "islem_tipi"      : tip,
        "banka_ref_no"    : m.group(1),
        "doviz_cinsi"     : "",
        "doviz_tutari"    : "",
        "kur"             : "",
        "karsitaraf_ad"   : "",
        "karsitaraf_iban" : "",
        "karsitaraf_banka": "Ziraat Katılım Bankası",
        "aciklama_temiz"  : acik,
        "cari_durumu"     : "BANKA_GİDERİ",
        "muhasebe_hesabi" : "780",
    }


def parse_IHMM(aciklama: str) -> dict | None:
    """
    FORMAT 3 — İhracat IBKB / TCMB Prim
    Örnek: 0145IHMM25000359 Ref. USD 183512 İhracat İBKB Alı - TCMB İhracat Karşılığı TCMB Kuru 40,005
           0145IHMM26000099 Ref. USD 250224 İhracat TCMB Döviz Dönüm Destekli Pirim Ödemesi
    """
    p = re.compile(
        r"(0145IHMM\S+)\s+Ref\.\s+([A-Z]{3})\s+([\d.,]+)\s+(.+)",
        re.IGNORECASE
    )
    m = p.search(aciklama)
    if not m:
        return None
    icerik = m.group(4).strip()
    # Kur var mı?
    kur_m = re.search(r"TCMB\s+Kuru\s+([\d.,]+)", icerik, re.IGNORECASE)
    kur   = kur_m.group(1).replace(",", ".") if kur_m else ""
    tip   = "TCMB_PRIM" if "PIRIM" in icerik.upper() or "PRİM" in icerik.upper() else "IHRACAT_IBKB"
    return {
        "islem_tipi"      : tip,
        "banka_ref_no"    : m.group(1),
        "doviz_cinsi"     : m.group(2).upper(),
        "doviz_tutari"    : m.group(3).replace(".", "").replace(",", "."),
        "kur"             : kur,
        "karsitaraf_ad"   : "TCMB",
        "karsitaraf_iban" : "",
        "karsitaraf_banka": "TCMB",
        "aciklama_temiz"  : icerik,
        "cari_durumu"     : "BANKA_GİDERİ",
        "muhasebe_hesabi" : "102",
    }


def parse_KRTM(aciklama: str) -> dict | None:
    """
    FORMAT 4 — Komisyon Gecikme Faizi ve BSMV
    Örnek (BSMV): 0145KRTM25000106 0,12 Komisyon Gecikme faizi BSMV TRY
    Örnek (Faiz): 0145KRTM25000106 237275,43 Komisyon Gecikme faizi 2,44
    """
    p = re.compile(r"(0145KRTM\S+)\s+([\d.,]+)\s+Komisyon\s+Gecikme\s+faizi\s+(.+)", re.IGNORECASE)
    m = p.search(aciklama)
    if not m:
        return None
    son_kisim = m.group(3).strip().upper()
    bsmv_mi   = "BSMV" in son_kisim
    tip       = "KOM_GEC_BSMV" if bsmv_mi else "KOM_GEC_FAIZ"
    doviz     = "TRY" if bsmv_mi else ""
    oran      = "" if bsmv_mi else son_kisim  # Faiz oranı
    return {
        "islem_tipi"      : tip,
        "banka_ref_no"    : m.group(1),
        "doviz_cinsi"     : doviz,
        "doviz_tutari"    : m.group(2).replace(",", "."),  # Faiz tutarı veya BSMV tutarı
        "kur"             : "",
        "karsitaraf_ad"   : "",
        "karsitaraf_iban" : "",
        "karsitaraf_banka": "Ziraat Katılım Bankası",
        "aciklama_temiz"  : f"Komisyon Gecikme Faizi {'BSMV' if bsmv_mi else ('Oran: %' + oran)}",
        "cari_durumu"     : "BANKA_GİDERİ",
        "muhasebe_hesabi" : "780",
    }


def parse_VADELI(aciklama: str) -> dict | None:
    """
    FORMAT 5 — Vadeli Hesap Virman / Kapama / Para Çekme
    Örnek: 145-97826392-5012 No.lu Vds. Hes. Aktararak 97826392-5066 No.lu Vadeli Hesap Kapama
           145-97826392-5012 No.lu Vds. Hes.tan 97826392-5041 -Hes.Açılı,Hes. Aç. Tr: ...
    """
    p = re.compile(
        r"(\d[\d\-]+)\s+No\.lu\s+Vds\.\s+Hes[\.\s]*(Aktararak|tan)\s+"
        r"(\d[\d\-]+)\s+No\.lu\s+(.+)",
        re.IGNORECASE
    )
    m = p.search(aciklama)
    if not m:
        return None
    kaynak    = m.group(1)
    hedef     = m.group(3)
    icerik    = m.group(4).strip()
    if "KAPAMA" in icerik.upper():
        tip   = "VADELI_KAPAMA"
        temiz = f"Vadeli Hesap Kapama — {kaynak} → {hedef}"
    elif "PARA ÇEKME" in icerik.upper() or "CEKME" in icerik.upper():
        tip   = "VADELI_PARA_CEKME"
        temiz = f"Vadeli Hesaptan Para Çekme — {kaynak} → {hedef}"
    else:
        tip   = "VADELI_AKTARIM"
        temiz = f"Vadeli Hesap Aktarım — {kaynak} → {hedef}"
    return {
        "islem_tipi"      : tip,
        "banka_ref_no"    : kaynak,
        "doviz_cinsi"     : "",
        "doviz_tutari"    : "",
        "kur"             : "",
        "karsitaraf_ad"   : "",
        "karsitaraf_iban" : "",
        "karsitaraf_banka": "Ziraat Katılım Bankası",
        "aciklama_temiz"  : temiz,
        "cari_durumu"     : "BANKA_GİDERİ",
        "muhasebe_hesabi" : "102",
    }


def parse_EFT_FAST(aciklama: str) -> dict | None:
    """
    FORMAT 6 — EFT veya FAST işlemi (QNB Bank üzerinden tedarikçi/personel)
    
    İki alt format:
    6a. QNB BANK A.Ş.//TR...--ALICI ADI/AÇIKLAMA
    6b. QNB Bank A.Ş./TR...-ALICI ADI/AÇIKLAMAFAST işlemi
    """
    # Format 6a: çift eğik çizgi ve çift tire
    p6a = re.compile(
        r"(QNB\s+BANK\s+A\.Ş\.)//\s*(TR\d{24})\s*--\s*([^/]+)/(.+)",
        re.IGNORECASE
    )
    # Format 6b: tekli ayraç, FAST sonda
    p6b = re.compile(
        r"(QNB\s+Bank\s+A\.Ş\.)/\s*(TR\d{24})\s*-\s*([^/]+)/(.+?)(?:FAST\s+i[şs]lemi)?$",
        re.IGNORECASE
    )
    fast = "FAST" in aciklama.upper()

    for p in [p6a, p6b]:
        m = p.search(aciklama)
        if m:
            alici_ad  = m.group(3).strip()
            acik_metn = m.group(4).strip().replace("FAST işlemi", "").strip()
            tip = "FAST_GIDEN" if fast else "EFT_GIDEN"
            return {
                "islem_tipi"      : tip,
                "banka_ref_no"    : "",
                "doviz_cinsi"     : "",
                "doviz_tutari"    : "",
                "kur"             : "",
                "karsitaraf_ad"   : alici_ad,
                "karsitaraf_iban" : iban_formatla(m.group(2)),
                "karsitaraf_banka": "QNB Bank A.Ş.",
                "aciklama_temiz"  : acik_metn,
                "cari_durumu"     : "CARİ_BELLİ",
                "muhasebe_hesabi" : "320",
            }
    return None


def parse_ZIRAAT_EFT(aciklama: str) -> dict | None:
    """
    FORMAT 7 — Ziraat Katılım veya diğer banka EFT (çift eğik çizgi + çift tire formatı)
    Örnek: ZIRAAT KATILIM BANKASI A.Ş.//TR...--FİRMA ADI/AÇIKLAMA
           Türkiye Garanti Bankası A.Ş./TR.../FİRMA ADI
           T.VAKIFLAR BANKASI T.A.O.//TR...--FİRMA ADI/AÇIKLAMA
    """
    # Çift eğik çizgi formatı: BANKA//IBAN--ALICI/ACIKLAMA
    p_cift = re.compile(
        r"([A-ZÇĞİÖŞÜa-zçğışöüñ\s\.]+A\.Ş\.|[A-ZÇĞİÖŞÜa-zçğışöüñ\s\.]+T\.A\.O\.|"
        r"[A-ZÇĞİÖŞÜa-zçğışöüñ\s\.]+A\.S\.)/+\s*(TR\d{24})\s*-{1,2}\s*([^/]*)/?(.*)",
        re.IGNORECASE
    )
    m = p_cift.search(aciklama)
    if not m:
        return None
    banka     = m.group(1).strip()
    iban      = m.group(2).strip()
    alici_ad  = m.group(3).strip().rstrip("-").strip()
    acik      = m.group(4).strip() if m.group(4) else ""
    # EFT mi HAVALE mi? "Mobil Havale" ifadesine bak
    tip = "HAVALE_GIDEN" if "HAVALE" in aciklama.upper() else "EFT_GIDEN"
    return {
        "islem_tipi"      : tip,
        "banka_ref_no"    : "",
        "doviz_cinsi"     : "",
        "doviz_tutari"    : "",
        "kur"             : "",
        "karsitaraf_ad"   : alici_ad,
        "karsitaraf_iban" : iban_formatla(iban),
        "karsitaraf_banka": banka,
        "aciklama_temiz"  : acik or alici_ad,
        "cari_durumu"     : "CARİ_BELLİ",
        "muhasebe_hesabi" : "320",
    }


def parse_ZIRAAT_HAVALE(aciklama: str) -> dict | None:
    """
    FORMAT 8 — Ziraat Mobil Havale (serbest metin formatı)
    Örnek: LEVARE ÖDEME FİRMA ADI FİRMA ADI TAM YAZIMI Ziraat Mobil Havale
           RFQ 1404 Remaining Payment FİRMA ADI Ziraat Mobil Havale
           Temmuz Ayı Kira Bedeli EGEÇELİ METAL ... Ziraat Mobil Havale
    Yapı: [AÇIKLAMA] [FİRMA KISA] [FİRMA TAM] Ziraat Mobil Havale
    """
    p = re.compile(r"(.+?)\s+Ziraat\s+Mobil\s+Havale\s*$", re.IGNORECASE)
    m = p.search(aciklama.strip())
    if not m:
        return None
    icerik = m.group(1).strip()
    # Son kısım genellikle firma adını iki kez tekrar ediyor,
    # ikinci tekrarı bul ve firma adı yap
    # Basit yaklaşım: son 60 karakterde firma adı olabilir
    return {
        "islem_tipi"      : "HAVALE_GIDEN",
        "banka_ref_no"    : "",
        "doviz_cinsi"     : "",
        "doviz_tutari"    : "",
        "kur"             : "",
        "karsitaraf_ad"   : _firma_adi_cikar(icerik),
        "karsitaraf_iban" : "",
        "karsitaraf_banka": "Ziraat Katılım Bankası",
        "aciklama_temiz"  : _aciklama_cikar(icerik),
        "cari_durumu"     : "CARİ_BELLİ",
        "muhasebe_hesabi" : "320",
    }


def _firma_adi_cikar(metin: str) -> str:
    """
    'AÇIKLAMA FİRMA_KISA FİRMA_TAM' formatından firma adını çıkarır.
    Firma adı genelde büyük harf ve Ltd/A.Ş. ile biter.
    """
    # Ltd.Şti, A.Ş., ANONİM ŞİRKETİ, LİMİTED ŞİRKETİ ile biten kısmı al
    m = re.search(
        r"([A-ZÇĞİÖŞÜ\s\.\-]+(?:A\.Ş\.|LTD\.\s*ŞTİ\.|ANONİM\s+ŞİRKETİ|LİMİTED\s+ŞİRKETİ|A\.S\.))\s*$",
        metin, re.IGNORECASE
    )
    if m:
        return m.group(1).strip()
    # Yoksa son 60 karakteri döndür
    return metin[-60:].strip() if len(metin) > 60 else metin


def _aciklama_cikar(metin: str) -> str:
    """Firma adından önceki açıklama kısmını çıkarır."""
    m = re.search(
        r"^(.+?)\s+[A-ZÇĞİÖŞÜ\s\.\-]+(?:A\.Ş\.|LTD\.|ANONİM|LİMİTED)",
        metin, re.IGNORECASE
    )
    return m.group(1).strip() if m else metin[:60].strip()


def parse_POS(aciklama: str) -> dict | None:
    """
    FORMAT 9 — POS Alışveriş (yurt içi, yurt dışı, sanal)
    Örnek: POS ALIŞVERİŞ (YURTİÇİ) KART NO: 5169 **** **** 6231 İŞYERİ ADI...
           SANAL POS ALIŞVERİŞ KART NO: 5169 **** **** 6231 ...
    """
    p = re.compile(
        r"((?:SANAL\s+)?POS\s+ALI[ŞS]VER[İI][ŞS](?:\s+\(YURT[İI]\s*Ç[İI]\s*\))?)\s+"
        r"KART\s+NO:\s*([\d\s\*]+)\s+(.*)",
        re.IGNORECASE
    )
    m = p.search(aciklama)
    if not m:
        return None
    tip      = "SANAL_POS" if "SANAL" in m.group(1).upper() else "POS_ALISVERIS"
    kart_no  = re.sub(r"\s+", " ", m.group(2).strip())
    isyeri   = m.group(3).strip()
    return {
        "islem_tipi"      : tip,
        "banka_ref_no"    : "",
        "doviz_cinsi"     : "",
        "doviz_tutari"    : "",
        "kur"             : "",
        "karsitaraf_ad"   : isyeri[:60],
        "karsitaraf_iban" : "",
        "karsitaraf_banka": "",
        "aciklama_temiz"  : f"POS Alışveriş — Kart: {kart_no} — {isyeri[:40]}",
        "cari_durumu"     : "BANKA_GİDERİ",
        "muhasebe_hesabi" : "760",
    }


def parse_BSMV_STANDALONE(aciklama: str) -> dict | None:
    """
    FORMAT 10 — Tek başına BSMV Tahsilatı
    Örnek: BSMV Tahsilatı
    """
    if re.match(r"^\s*BSMV\s+Tahsilat[ıi]\s*$", aciklama, re.IGNORECASE):
        return {
            "islem_tipi"      : "BSMV",
            "banka_ref_no"    : "",
            "doviz_cinsi"     : "",
            "doviz_tutari"    : "",
            "kur"             : "",
            "karsitaraf_ad"   : "",
            "karsitaraf_iban" : "",
            "karsitaraf_banka": "Ziraat Katılım Bankası",
            "aciklama_temiz"  : "BSMV Tahsilatı",
            "cari_durumu"     : "BANKA_GİDERİ",
            "muhasebe_hesabi" : "780",
        }
    return None


def parse_GONDEREN(aciklama: str) -> dict | None:
    """
    FORMAT 11 — Gönderen (Gnd: / Gönderen:) formatı
    Örnek: Gnd: ALLİANZ SİGORTA A.Ş. Ref:S20251202...
    """
    p = re.compile(r"Gnd(?:eren)?:\s*(.+?)(?:\s+Ref:(.+))?$", re.IGNORECASE)
    m = p.search(aciklama)
    if not m:
        return None
    gonderen = m.group(1).strip()
    ref      = m.group(2).strip() if m.group(2) else ""
    return {
        "islem_tipi"      : "EFT_GELEN",
        "banka_ref_no"    : ref,
        "doviz_cinsi"     : "",
        "doviz_tutari"    : "",
        "kur"             : "",
        "karsitaraf_ad"   : gonderen,
        "karsitaraf_iban" : "",
        "karsitaraf_banka": "",
        "aciklama_temiz"  : f"Gelen ödeme — {gonderen}",
        "cari_durumu"     : "CARİ_BELLİ",
        "muhasebe_hesabi" : "120",
    }


def parse_FATURA_ODEME(aciklama: str) -> dict | None:
    """
    FORMAT 12 — Fatura ödemeleri (Turkcell, Telekom vb.)
    Örnek: TURKCELL/Tel No: 5346961480 /Fatura No: ...
           Telekom Online tah/ 2011610445394 / ...
    """
    p = re.compile(r"(TURKCELL|TELEKOM|TÜRKTELEKOM|VODAFONE|TÜRK TELEKOM)\s*/(.+)", re.IGNORECASE)
    m = p.search(aciklama)
    if not m:
        return None
    return {
        "islem_tipi"      : "FATURA_ODEME",
        "banka_ref_no"    : "",
        "doviz_cinsi"     : "",
        "doviz_tutari"    : "",
        "kur"             : "",
        "karsitaraf_ad"   : m.group(1).upper(),
        "karsitaraf_iban" : "",
        "karsitaraf_banka": "",
        "aciklama_temiz"  : m.group(2).strip()[:80],
        "cari_durumu"     : "BANKA_GİDERİ",
        "muhasebe_hesabi" : "770",
    }


def parse_VERGI(aciklama: str) -> dict | None:
    """
    FORMAT 13 — Vergi ödemeleri
    Örnek: VERGI-4017 KATMA DEGER VER VD:035250-...
           VERGI-0003 GELİR VERGİSİ ...
    """
    p = re.compile(r"VERG[İI]-(\d+)\s+(.+?)(?:\s+VD:(\S+))?(?:\s+VN:(\S+))?", re.IGNORECASE)
    m = p.search(aciklama)
    if not m:
        return None
    return {
        "islem_tipi"      : "VERGI",
        "banka_ref_no"    : m.group(1),
        "doviz_cinsi"     : "",
        "doviz_tutari"    : "",
        "kur"             : "",
        "karsitaraf_ad"   : "Gelir İdaresi Başkanlığı",
        "karsitaraf_iban" : "",
        "karsitaraf_banka": "",
        "aciklama_temiz"  : m.group(2).strip()[:80],
        "cari_durumu"     : "DEVLET_ÖDEMESİ",
        "muhasebe_hesabi" : "360",
    }


def parse_MESAI_GUMRUK(aciklama: str) -> dict | None:
    """
    FORMAT 14 — Mesai / Gümrük ödemeleri (Hazine)
    Örnek: MESAI - - 26341453EX00042897 // 6081724158 LEVARE VSD... HAZ.İÇ ÖDEM.TAHSİLAT
    """
    p = re.compile(r"MESA[İI](.+?)(?:HAZ\.?\s*[İI][Çç]|HAZINE)", re.IGNORECASE)
    m = p.search(aciklama)
    if not m:
        return None
    return {
        "islem_tipi"      : "GUMRUK_MESAI",
        "banka_ref_no"    : "",
        "doviz_cinsi"     : "",
        "doviz_tutari"    : "",
        "kur"             : "",
        "karsitaraf_ad"   : "T.C. Hazine",
        "karsitaraf_iban" : "",
        "karsitaraf_banka": "",
        "aciklama_temiz"  : aciklama[:100].strip(),
        "cari_durumu"     : "DEVLET_ÖDEMESİ",
        "muhasebe_hesabi" : "360",
    }


def parse_IHPE(aciklama: str) -> dict | None:
    """FORMAT — İhracat IHPE (IHMM'e benzer farklı prefix)"""
    p = re.compile(
        r"(0145IHPE\S+)\s+Ref\.\s+([A-Z]{3})\s+([\d.,]+)\s+(.+)",
        re.IGNORECASE
    )
    m = p.search(aciklama)
    if not m:
        return None
    kur_m = re.search(r"TCMB\s+Kuru\s+([\d.,]+)", m.group(4), re.IGNORECASE)
    return {
        "islem_tipi"      : "IHRACAT_IBKB",
        "banka_ref_no"    : m.group(1),
        "doviz_cinsi"     : m.group(2).upper(),
        "doviz_tutari"    : m.group(3).replace(".", "").replace(",", "."),
        "kur"             : kur_m.group(1).replace(",", ".") if kur_m else "",
        "karsitaraf_ad"   : "TCMB",
        "karsitaraf_iban" : "",
        "karsitaraf_banka": "TCMB",
        "aciklama_temiz"  : m.group(4).strip()[:80],
        "cari_durumu"     : "BANKA_GİDERİ",
        "muhasebe_hesabi" : "102",
    }


def parse_BSMV_KOMISYON_TUTARI(aciklama: str) -> dict | None:
    """
    FORMAT — BSMV TUTARI / KOMİSYON TUTARI (tek başına)
    Örnek: BSMV TUTARI
           KOMİSYON TUTARI
    """
    if re.match(r"^\s*(BSMV\s+TUTARI|KOM[İI]SYON\s+TUTARI)\s*$", aciklama, re.IGNORECASE):
        tip = "BSMV" if "BSMV" in aciklama.upper() else "KOMISYON"
        return {
            "islem_tipi"      : tip,
            "banka_ref_no"    : "",
            "doviz_cinsi"     : "",
            "doviz_tutari"    : "",
            "kur"             : "",
            "karsitaraf_ad"   : "",
            "karsitaraf_iban" : "",
            "karsitaraf_banka": "Ziraat Katılım Bankası",
            "aciklama_temiz"  : aciklama.strip(),
            "cari_durumu"     : "BANKA_GİDERİ",
            "muhasebe_hesabi" : "780",
        }
    return None


def parse_HGS_OGS(aciklama: str) -> dict | None:
    """
    FORMAT — Otomatik HGS/OGS geçiş ücreti
    Örnek: OTOMATIK Plaka no: 35CLF059 , 1117171809 nolu...
    """
    p = re.compile(r"OTOMATIK\s+Plaka\s+no:\s*(\S+)", re.IGNORECASE)
    m = p.search(aciklama)
    if not m:
        return None
    return {
        "islem_tipi"      : "HGS_OGS",
        "banka_ref_no"    : "",
        "doviz_cinsi"     : "",
        "doviz_tutari"    : "",
        "kur"             : "",
        "karsitaraf_ad"   : "KGM / HGS",
        "karsitaraf_iban" : "",
        "karsitaraf_banka": "",
        "aciklama_temiz"  : f"Otomatik Geçiş — Plaka: {m.group(1)}",
        "cari_durumu"     : "BANKA_GİDERİ",
        "muhasebe_hesabi" : "770",
    }


def parse_SGK(aciklama: str) -> dict | None:
    """
    FORMAT — SGK Ödemesi
    Örnek: SGK( EMS/BAĞ/SSK)/ 22712010118819850351032000...
    """
    if re.search(r"SGK\s*[\(/]", aciklama, re.IGNORECASE):
        return {
            "islem_tipi"      : "SGK",
            "banka_ref_no"    : "",
            "doviz_cinsi"     : "",
            "doviz_tutari"    : "",
            "kur"             : "",
            "karsitaraf_ad"   : "SGK",
            "karsitaraf_iban" : "",
            "karsitaraf_banka": "",
            "aciklama_temiz"  : aciklama[:80].strip(),
            "cari_durumu"     : "DEVLET_ÖDEMESİ",
            "muhasebe_hesabi" : "361",
        }
    return None


def parse_LIMIT_MASRAF(aciklama: str) -> dict | None:
    """
    FORMAT — Limit Tahsisi / Kredi İşlem Masrafı
    Örnek: LİMİT TAHSİSİ - Kredi İşlem Masrafı
           iade fişi LİMİT TAHSİSİ - Kredi İşlem Masrafı
    """
    if re.search(r"L[İI]M[İI]T\s+TAHS[İI]S", aciklama, re.IGNORECASE):
        return {
            "islem_tipi"      : "KREDI_MASRAF",
            "banka_ref_no"    : "",
            "doviz_cinsi"     : "",
            "doviz_tutari"    : "",
            "kur"             : "",
            "karsitaraf_ad"   : "",
            "karsitaraf_iban" : "",
            "karsitaraf_banka": "Ziraat Katılım Bankası",
            "aciklama_temiz"  : aciklama[:80].strip(),
            "cari_durumu"     : "BANKA_GİDERİ",
            "muhasebe_hesabi" : "780",
        }
    return None


def parse_SIPARIS_KODU(aciklama: str) -> dict | None:
    """
    FORMAT — Sipariş kodu + firma + Ziraat Mobil Havale (banka IBAN içermeyen)
    Örnek: RFQ 1239 SC12025000013565 SABİT TEKNİK HIRDAVAT... Ziraat Mobil Havale
           BLM2025000004603 BOLEM ELEKTRİK... Ziraat Mobil Havale
           AEF2025000001740 İŞYERİ HEKİMİ ÜCRETİ ATLAS...
    
    Genel kural: Büyük harf kodu ile başlayan ve firma adı içeren işlemler
    """
    # Ziraat Mobil Havale zaten parse_ZIRAAT_HAVALE tarafından yakalanıyor
    # Buraya sadece Havale kelimesi olmayan ama kod+firma formatındakiler düşer
    p = re.compile(
        r"^([A-Z0-9]{3,}\d{9,})\s+(.+?)\s+([A-ZÇĞİÖŞÜ\s\.\-]+(?:A\.Ş\.|LTD|ANONİM|LİMİTED|A\.S\.).*?)$",
        re.IGNORECASE
    )
    m = p.match(aciklama.strip())
    if not m:
        return None
    return {
        "islem_tipi"      : "HAVALE_GIDEN",
        "banka_ref_no"    : m.group(1),
        "doviz_cinsi"     : "",
        "doviz_tutari"    : "",
        "kur"             : "",
        "karsitaraf_ad"   : m.group(3).strip()[:80],
        "karsitaraf_iban" : "",
        "karsitaraf_banka": "Ziraat Katılım Bankası",
        "aciklama_temiz"  : m.group(2).strip()[:80],
        "cari_durumu"     : "CARİ_BELLİ",
        "muhasebe_hesabi" : "320",
    }


def parse_VADELI_HESAP_NO(aciklama: str) -> dict | None:
    """
    FORMAT 15 — Sadece hesap numarasıyla başlayan vadeli işlemler
    Örnek: 97826392-5078 (Vadeli hesap hareketleri)
    """
    p = re.compile(r"^(\d{8}-\d{4})\s*$")
    m = p.match(aciklama.strip())
    if not m:
        return None
    return {
        "islem_tipi"      : "VADELI_AKTARIM",
        "banka_ref_no"    : m.group(1),
        "doviz_cinsi"     : "",
        "doviz_tutari"    : "",
        "kur"             : "",
        "karsitaraf_ad"   : "",
        "karsitaraf_iban" : "",
        "karsitaraf_banka": "Ziraat Katılım Bankası",
        "aciklama_temiz"  : f"Vadeli Hesap: {m.group(1)}",
        "cari_durumu"     : "BANKA_GİDERİ",
        "muhasebe_hesabi" : "102",
    }


# ─────────────────────────────────────────
# 4. ANA PARSER (tüm formatları sırayla dener)
# ─────────────────────────────────────────
PARSER_SIRASI = [
    parse_BSMV_STANDALONE,       # Tam eşleşme: "BSMV Tahsilatı"
    parse_BSMV_KOMISYON_TUTARI,  # Tam eşleşme: "BSMV TUTARI" / "KOMİSYON TUTARI"
    parse_VADELI_HESAP_NO,       # Tam eşleşme: "97826392-5078"
    parse_LIMIT_MASRAF,          # Anahtar kelime: "LİMİT TAHSİSİ"
    parse_HGS_OGS,               # Anahtar kelime: "OTOMATIK Plaka"
    parse_SGK,                   # Anahtar kelime: "SGK("
    parse_VERGI,                 # Anahtar kelime: "VERGI-XXXX"
    parse_MESAI_GUMRUK,          # Anahtar kelime: "MESAI...HAZ"
    parse_FATURA_ODEME,          # Anahtar kelime: "TURKCELL/" vb.
    parse_GONDEREN,              # "Gnd:"
    parse_GKDA,                  # "0145GKDA"
    parse_HOHH,                  # "0145HOHH"
    parse_IHMM,                  # "0145IHMM"
    parse_IHPE,                  # "0145IHPE"
    parse_KRTM,                  # "0145KRTM"
    parse_VADELI,                # "No.lu Vds. Hes"
    parse_EFT_FAST,              # "QNB BANK" EFT/FAST
    parse_ZIRAAT_EFT,            # Banka adı + IBAN formatı
    parse_ZIRAAT_HAVALE,         # "... Ziraat Mobil Havale" (serbest metin)
    parse_POS,                   # "POS ALIŞVERİŞ"
    parse_SIPARIS_KODU,          # Sipariş kodu + firma adı
]

BOSH_SONUC = {
    "islem_tipi"      : "BELİRSİZ",
    "banka_ref_no"    : "",
    "doviz_cinsi"     : "",
    "doviz_tutari"    : "",
    "kur"             : "",
    "karsitaraf_ad"   : "",
    "karsitaraf_iban" : "",
    "karsitaraf_banka": "",
    "aciklama_temiz"  : "",
    "cari_durumu"     : "BELİRSİZ",
    "muhasebe_hesabi" : "",
    "confidence_score": 0,
    "explanation"     : ""
}

def satir_parse_et(aciklama: str, islem_yonu: str) -> dict:
    # 1. Önce Hızlı & Ücretsiz Regex Parser'ları dene
    for parser in PARSER_SIRASI:
        sonuc = parser(aciklama)
        if sonuc:
            tam_sonuc = BOSH_SONUC.copy()
            tam_sonuc.update(sonuc)
            # Regex tabanlı kurallar her zaman %100 güvenilirdir
            tam_sonuc["confidence_score"] = 100
            tam_sonuc["explanation"] = "Banka formatı regex ile kesin olarak tanındı."
            return tam_sonuc
            
    # 2. Eğer Regex ile çözülemezse Yapay Zeka (Gemini Agent) devralır
    try:
        from ai_agent import parse_with_ai
        ai_sonuc = parse_with_ai(aciklama, islem_yonu)
        if ai_sonuc:
            tam_sonuc = BOSH_SONUC.copy()
            tam_sonuc.update(ai_sonuc)
            return tam_sonuc
    except Exception as e:
        print(f"AI Parse Hatası: {e}")
        
    # Her şeye rağmen anlaşılamazsa BOSH_SONUC dön
    return BOSH_SONUC.copy()


# ─────────────────────────────────────────
# 5. TOPLU PARSE + ÇIKTI
# ─────────────────────────────────────────
def main():
    print(f"Dosya okunuyor: {GIRIS_DOSYA}")
    df = veri_oku(GIRIS_DOSYA)
    print(f"Toplam satır: {len(df)}")

    # Her satırı parse et
    sonuclar = []
    from tqdm import tqdm
    print("Satırlar analiz ediliyor...")
    for i, row in tqdm(df.iterrows(), total=len(df)):
        parsed = satir_parse_et(str(row["aciklama_ham"]), str(row["islem_yonu"]))
        sonuclar.append(parsed)

    df_parsed = pd.DataFrame(sonuclar)

    # Orijinal + parse edilmiş sütunları birleştir
    df_final = pd.concat([
        df[["islem_tarihi", "fis_no", "islem_yonu", "aciklama_ham"]].reset_index(drop=True),
        df_parsed.reset_index(drop=True)
    ], axis=1)

    # İstatistik
    tipler     = df_final["islem_tipi"].value_counts()
    belirsizler = (df_final["islem_tipi"] == "BELİRSİZ").sum()
    print("\n=== PARSE SONUCU ===")
    print(tipler.to_string())
    print(f"\nBELİRSİZ (parse edilemeyen): {belirsizler} / {len(df_final)}")
    print(f"Başarı oranı: %{100 * (1 - belirsizler/len(df_final)):.1f}")


    # Excel'e yaz
    with pd.ExcelWriter(CIKIS_DOSYA, engine="openpyxl") as writer:
        df_final.to_excel(writer, index=False, sheet_name="Parsed")

        # Renk kodlama
        from openpyxl.styles import PatternFill, Font
        ws = writer.sheets["Parsed"]

        RENKLER = {
            "BANKA_GİDERİ" : "D9EAD3",  # Açık yeşil
            "CARİ_BELLİ"   : "CFE2F3",  # Açık mavi
            "BELİRSİZ"      : "FCE5CD",  # Turuncu
        }

        # Başlık satırı koyu
        for cell in ws[1]:
            cell.font = Font(bold=True)

        # Cari durumuna göre satır rengi
        cari_col_idx = list(df_final.columns).index("cari_durumu") + 1
        for row_idx in range(2, len(df_final) + 2):
            cari_val = ws.cell(row=row_idx, column=cari_col_idx).value or ""
            renk     = RENKLER.get(cari_val, "FFFFFF")
            fill     = PatternFill(fill_type="solid", fgColor=renk)
            for cell in ws[row_idx]:
                cell.fill = fill

        # Sütun genişlikleri
        for col in ws.columns:
            max_len = max(len(str(c.value or "")) for c in col)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 50)

    print(f"\nÇıktı kaydedildi: {CIKIS_DOSYA}")


if __name__ == "__main__":
    main()
